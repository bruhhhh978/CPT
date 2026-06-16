import calendar
import datetime
from urllib import request
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.db.models import Sum, Q, Count
from .models import Employee, Attendance, Adjustment, AdjustmentLog, CongTrinh, NhanVien, DanhMucNghe, ChamCongHangNgay, PhuThuThuongPhat, ChotLuongThang
from decimal import Decimal
import pandas as pd
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from django.contrib.humanize.templatetags.humanize import intcomma
from django.template.loader import render_to_string
from django.contrib.auth.models import User
from login.models import UserProfile,  UserLoginLog, UserActivityLog, ActiveSession, log_activity
from .decorators import manager_only, user_and_manager, allow_viewer
from django.contrib import messages
from .forms import CongTrinhForm
from datetime import timedelta
from django.db.models.functions import TruncWeek, TruncMonth, TruncYear
import datetime
import calendar
from decimal import Decimal, ROUND_HALF_UP
from django.contrib.auth.decorators import login_required

# ============ MANAGER DASHBOARD & USER MANAGEMENT ============

@manager_only
def manager_dashboard(request):
    search_query = request.GET.get('q', '').strip()
    users = User.objects.all().select_related('profile')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    context = {
        'users': users,
        'user_count': users.count(),
        'manager_count': users.filter(profile__role='manager').count(),
        'user_count_role': users.filter(profile__role='user').count(),
        'viewer_count': users.filter(profile__role='viewer').count(),
        'search_query': search_query,
    }
    return render(request, 'payroll/manager_dashboard.html', context)

@manager_only
def create_user(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        email = request.POST.get('email', '').strip()
        role = request.POST.get('role', 'user')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        if not username or len(username) < 3:
            messages.error(request, 'Tên đăng nhập phải ít nhất 3 ký tự.')
            return redirect('payroll:manager_dashboard')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Tên đăng nhập đã tồn tại.')
            return redirect('payroll:manager_dashboard')
        if not password:
            messages.error(request, 'Mật khẩu không được để trống.')
            return redirect('payroll:manager_dashboard')
        if len(password) < 6:
            messages.error(request, 'Mật khẩu không đúng quy định: Phải có ít nhất 6 ký tự.')
            return redirect('payroll:manager_dashboard')
        if role not in ['manager', 'user', 'viewer']:
            messages.error(request, 'Cấp độ không hợp lệ.')
            return redirect('payroll:manager_dashboard')
        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username, password=password, email=email,
                    first_name=first_name, last_name=last_name
                )
                UserProfile.objects.update_or_create(user=user, defaults={'role': role})
            messages.success(request, f'Tạo tài khoản "{username}" với cấp "{role}" thành công.')
            log_activity(request, 'CREATE', 'TaiKhoan', object_repr=username, description=f'Tạo tài khoản: {username} (cấp {role})')
        except Exception as e:
            messages.error(request, f'Lỗi khi tạo tài khoản: {str(e)}')
        return redirect('payroll:manager_dashboard')
    return redirect('payroll:manager_dashboard')

@manager_only
def edit_user(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            if 'first_name' in request.POST:
                user.first_name = request.POST.get('first_name', '').strip()
            if 'last_name' in request.POST:
                user.last_name = request.POST.get('last_name', '').strip()
            if 'email' in request.POST:
                user.email = request.POST.get('email', '').strip()
            user.save()
            new_role = request.POST.get('role')
            if new_role:
                if user == request.user and new_role != 'manager':
                    messages.error(request, 'Bạn không thể hạ cấp chính mình.')
                else:
                    user.profile.role = new_role
                    user.profile.save()
            messages.success(request, f'Cập nhật thông tin cho "{user.username}" thành công.')
            log_activity(request, 'UPDATE', 'TaiKhoan', object_id=str(user_id), object_repr=user.username, description=f'Cập nhật tài khoản: {user.username}')
        except User.DoesNotExist:
            messages.error(request, 'Người dùng không tồn tại.')
        except Exception as e:
            messages.error(request, f'Lỗi khi cập nhật: {str(e)}')
    return redirect('payroll:manager_dashboard')

@manager_only
def toggle_user_status(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            if user == request.user:
                messages.error(request, 'Bạn không thể tự khóa tài khoản của chính mình.')
            else:
                user.is_active = not user.is_active
                user.save()
                status_msg = "Mở khóa" if user.is_active else "Khóa"
                messages.success(request, f'{status_msg} tài khoản "{user.username}" thành công.')
                log_activity(request, 'UPDATE', 'TaiKhoan', object_repr=user.username, description=f'{status_msg} tài khoản: {user.username}')
        except User.DoesNotExist:
            messages.error(request, 'Người dùng không tồn tại.')
    return redirect('payroll:manager_dashboard')

@manager_only
def reset_user_password(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            new_password = request.POST.get('new_password', '').strip()
            if not new_password or len(new_password) < 6:
                messages.error(request, f'Lỗi: Mật khẩu cho tài khoản "{user.username}" phải từ 6 ký tự trở lên.')
            else:
                user.set_password(new_password)
                user.save()
                messages.success(request, f'Đã đặt lại mật khẩu cho tài khoản "{user.username}".')
                log_activity(request, 'UPDATE', 'TaiKhoan', object_repr=user.username, description=f'Đặt lại mật khẩu: {user.username}')
        except User.DoesNotExist:
            messages.error(request, 'Người dùng không tồn tại.')
        except Exception as e:
            messages.error(request, f'Lỗi: {str(e)}')
    return redirect('payroll:manager_dashboard')

@manager_only
def delete_user(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            if user == request.user:
                messages.error(request, 'Bạn không thể xóa chính mình.')
                return redirect('payroll:manager_dashboard')
            username = user.username
            user.delete()
            log_activity(request, 'DELETE', 'TaiKhoan', object_repr=username, description=f'Xóa tài khoản: {username}')
            messages.success(request, f'Xóa tài khoản "{username}" thành công.')
        except User.DoesNotExist:
            messages.error(request, 'Người dùng không tồn tại.')
        except Exception as e:
            messages.error(request, f'Lỗi khi xóa: {str(e)}')
    return redirect('payroll:manager_dashboard')

@user_and_manager
def change_own_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        if not request.user.check_password(old_password):
            messages.error(request, 'Mật khẩu cũ không chính xác.')
        elif new_password != confirm_password:
            messages.error(request, 'Mật khẩu xác nhận không khớp.')
        elif len(new_password) < 6:
            messages.error(request, 'Mật khẩu mới phải có ít nhất 6 ký tự.')
        else:
            request.user.set_password(new_password)
            request.user.save()
            messages.success(request, 'Đổi mật khẩu thành công. Vui lòng đăng nhập lại.')
            return redirect('login:login')
    return render(request, 'payroll/change_password.html')

# ============ END MANAGER DASHBOARD ============

def get_week_start(date):
    days_since_friday = (date.weekday() - 4) % 7
    return date - datetime.timedelta(days=days_since_friday)

def get_week_range(date):
    friday = get_week_start(date)
    return [friday + datetime.timedelta(days=i) for i in range(7)]


# ============================================================
# ✅ FIX: get_available_weeks — sinh tuần từ khoảng thời gian
#         dự án thay vì chỉ lấy tuần có Attendance record
# ============================================================
def get_available_weeks(du_an_id=None):
    from .models import Attendance, CongTrinh
    weeks = set()

    # Nếu có dự án → sinh tất cả các tuần từ ngày bắt đầu đến ngày kết thúc
    if du_an_id:
        try:
            du_an = CongTrinh.objects.get(id=du_an_id)
            start = du_an.ngay_bat_dau
            end = du_an.thoi_han_ket_thuc or datetime.date.today()

            # Bắt đầu từ thứ 6 đầu tiên chứa ngày bắt đầu dự án
            current = get_week_start(start)
            while current <= end:
                weeks.add(current)
                current += datetime.timedelta(days=7)
        except (CongTrinh.DoesNotExist, ValueError):
            pass

    # Luôn bổ sung thêm các tuần có dữ liệu thực tế
    # (để không mất data nếu có attendance nằm ngoài khoảng dự án)
    qs = Attendance.objects.all()
    if du_an_id:
        qs = qs.filter(cong_trinh_id=du_an_id)
    for d in qs.values_list('date', flat=True).distinct():
        weeks.add(get_week_start(d))

    # Fallback: nếu không có dự án và không có data → dùng tuần hiện tại
    if not weeks:
        weeks.add(get_week_start(datetime.date.today()))

    return sorted(weeks)


# ============================================================
# ✅ FIX: get_available_months — sinh tháng từ khoảng thời
#         gian dự án thay vì chỉ từ Attendance
# ============================================================
def get_available_months(du_an_id=None):
    from .models import Attendance, CongTrinh
    months = set()

    if du_an_id:
        try:
            du_an = CongTrinh.objects.get(id=du_an_id)
            start = du_an.ngay_bat_dau
            end = du_an.thoi_han_ket_thuc or datetime.date.today()

            current = get_week_start(start)
            while current <= end:
                months.add((current.year, current.month))
                current += datetime.timedelta(days=7)
        except (CongTrinh.DoesNotExist, ValueError):
            pass

    qs = Attendance.objects.all()
    if du_an_id:
        qs = qs.filter(cong_trinh_id=du_an_id)
    for d in qs.values_list('date', flat=True).distinct():
        friday = get_week_start(d)
        months.add((friday.year, friday.month))

    return sorted(months)


def get_weeks_in_month(year, month, du_an_id=None):
    all_weeks = get_available_weeks(du_an_id=du_an_id)
    return [w for w in all_weeks if w.year == year and w.month == month]

def to_symbol(val):
    if val >= 1.0: return 'x'
    if val >= 0.5: return '/'
    return ''

def from_symbol(val):
    if val is None: return Decimal('0.0')
    val_str = str(val).strip().lower()
    if val_str == 'x': return Decimal('1.0')
    if val_str in ('/', '\\'): return Decimal('0.5')
    if not val_str: return Decimal('0.0')
    try:
        return Decimal(val_str)
    except:
        return Decimal('0.0')

def get_target_date(request):
    date_str = request.GET.get('date')
    if not date_str or date_str == "None":
        return timezone.now().date()
    try:
        return datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return timezone.now().date()

def get_weekday_label(date):
    lookup = {0: 'T2', 1: 'T3', 2: 'T4', 3: 'T5', 4: 'T6', 5: 'T7', 6: 'CN'}
    return lookup.get(date.weekday(), date.strftime('%a'))

def parse_attendance_input(value):
    if value is None:
        return Decimal('0.0')
    value_str = str(value).strip()
    if not value_str:
        return Decimal('0.0')
    try:
        return from_symbol(value_str)
    except Exception:
        try:
            return Decimal(value_str.replace(',', '.'))
        except Exception:
            return Decimal('0.0')

def get_search_type(request):
    search_type = request.GET.get('search_type', 'name').strip().lower()
    return search_type if search_type in ('name', 'dob') else 'name'

def get_filtered_employees(search_query='', search_type='name'):
    employees = Employee.objects.all().order_by('name')
    if search_query:
        employees = employees.filter(build_employee_search_query(search_query, search_type))
    return employees

def parse_birth_date_query(search_query):
    normalized_query = search_query.strip()
    for date_format in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y'):
        try:
            return {'match_type': 'full_date', 'value': datetime.datetime.strptime(normalized_query, date_format).date()}
        except ValueError:
            continue
    for date_format in ('%d/%m', '%d-%m'):
        try:
            parsed_date = datetime.datetime.strptime(normalized_query, date_format)
            return {'match_type': 'day_month', 'day': parsed_date.day, 'month': parsed_date.month}
        except ValueError:
            continue
    return None

def build_employee_search_query(search_query, search_type='name'):
    if search_type == 'dob':
        birth_date_query = parse_birth_date_query(search_query)
        if not birth_date_query:
            return Q(pk__in=[])
        if birth_date_query['match_type'] == 'full_date':
            return Q(date_of_birth=birth_date_query['value'])
        return Q(date_of_birth__day=birth_date_query['day'], date_of_birth__month=birth_date_query['month'])
    return Q(name__icontains=search_query)

def build_weekly_payroll_data(employees, dates, du_an_id=None):
    start_date, end_date = dates[0], dates[-1]
    payroll_data = []
    for emp in employees:
        att_qs = Attendance.objects.filter(employee=emp, date__range=(start_date, end_date))
        if du_an_id:
            att_qs = att_qs.filter(cong_trinh_id=du_an_id)
        attendance_map = {att.date: att for att in att_qs}
        daily_atts = []
        total_hc = Decimal('0.0')
        total_tc = Decimal('0.0')
        for d in dates:
            att = attendance_map.get(d)
            hc = att.regular_workday if att else Decimal('0.0')
            tc = att.overtime_workday if att else Decimal('0.0')
            total_hc += hc
            total_tc += tc
            daily_atts.append({
                'date': d,
                'hc_symbol': to_symbol(hc) if hc > 0 else '',
                'tc_symbol': str(tc) if tc > 0 else ''
            })
        adj = Adjustment.objects.filter(employee=emp, start_date=start_date, end_date=end_date).first()
        adjustment_val = adj.amount if adj else Decimal('0.0')
        total_pay = (total_hc * emp.daily_wage) + (total_tc * (emp.daily_wage / Decimal('8'))) + adjustment_val
        payroll_data.append({
            'employee': emp,
            'daily_attendance': daily_atts,
            'total_hc': total_hc.quantize(Decimal('1'), rounding=ROUND_HALF_UP),
            'total_tc': total_tc.quantize(Decimal('1'), rounding=ROUND_HALF_UP),
            'adjustment': adjustment_val,
            'total_pay': total_pay
        })
    return payroll_data


def payroll_sheet(request, pk=None):
    du_an_id = pk or request.GET.get('du_an_id') or ''
    du_an_id = str(du_an_id).strip() if du_an_id else ''
    du_an_obj = None
    if du_an_id:
        try:
            du_an_obj = CongTrinh.objects.get(id=du_an_id)
        except (CongTrinh.DoesNotExist, ValueError):
            du_an_id = ''

    can_edit = False
    is_viewer_mode = True
    if request.user.is_authenticated:
        try:
            role = request.user.profile.role
            if role in ['manager', 'user']:
                can_edit = True
                is_viewer_mode = False
        except:
            messages.error(request, 'Lỗi: Hồ sơ người dùng không hợp lệ.')
            return redirect('login:login')

    view_type = request.GET.get('view_type', 'week').strip().lower()
    if view_type not in ('week', 'month'):
        view_type = 'week'

    # ✅ FIX: Nếu có dự án và không truyền ?date= → tự động về tuần đầu dự án
    if du_an_obj and 'date' not in request.GET:
        target_date = get_week_start(du_an_obj.ngay_bat_dau)
    else:
        target_date = get_target_date(request)

    search_query = request.GET.get('q', '').strip()
    SUMMARY_NAMES = [
        'Tổng số Cai', 'Tổng số Kho', 'Tổng số LĐ', 
        'Người chấm công', 'Cộng', 'Tổng cộng'
    ]

    if du_an_id:
        emp_ids_in_project = Attendance.objects.filter(
            cong_trinh_id=du_an_id
        ).values_list('employee_id', flat=True).distinct()
        employees = Employee.objects.filter(
            id__in=emp_ids_in_project
        ).exclude(name__in=SUMMARY_NAMES).order_by('name')
    else:
        employees = Employee.objects.exclude(name__in=SUMMARY_NAMES).order_by('name')

    if search_query:
        employees = employees.filter(name__icontains=search_query)

    search_type = get_search_type(request)

    if view_type == 'month':
        week_start = get_week_start(target_date)
        year, month = week_start.year, week_start.month
        weeks_this_month = get_weeks_in_month(year, month, du_an_id=du_an_id)
        if weeks_this_month:
            first_day = weeks_this_month[0]
            last_day = weeks_this_month[-1] + datetime.timedelta(days=6)
            dates = [first_day + datetime.timedelta(days=i) for i in range((last_day - first_day).days + 1)]
        else:
            dates = get_week_range(target_date)
    else:
        dates = get_week_range(target_date)

    start_date, end_date = dates[0], dates[-1]

    if request.GET.get('export') == 'excel':
        if not request.user.is_authenticated:
            messages.error(request, 'Vui lòng đăng nhập để xuất Excel.')
            return redirect('login:login')

        log_activity(request, 'EXPORT', 'BaoCao',
                 description=f'Xuất Excel bảng lương {start_date.strftime("%d/%m")} - {end_date.strftime("%d/%m/%Y")}')

        return export_payroll_excel(dates, start_date, end_date, view_type, du_an_id=du_an_id)

    payroll_data = build_weekly_payroll_data(employees, dates, du_an_id=du_an_id)

    # Danh sách tuần — lấy từ khoảng thời gian dự án (đã fix)
    all_weeks_raw = get_available_weeks(du_an_id=du_an_id)
    available_weeks = [
        {
            'start': w,
            'end': w + datetime.timedelta(days=6),
            'date_str': w.strftime('%Y-%m-%d'),
            'label': f"Tuần {i:02d}: {w.strftime('%d/%m')} – {(w + datetime.timedelta(days=6)).strftime('%d/%m/%Y')}",
        }
        for i, w in enumerate(all_weeks_raw, start=1)
    ]

    # Danh sách tháng — lấy từ khoảng thời gian dự án (đã fix)
    available_months = []
    for (year, month) in get_available_months(du_an_id=du_an_id):
        weeks_in_m = get_weeks_in_month(year, month, du_an_id=du_an_id)
        first_week_str = weeks_in_m[0].strftime('%Y-%m-%d') if weeks_in_m else f'{year}-{month:02d}-01'
        available_months.append({
            'year': year, 'month': month, 'date_str': first_week_str,
            'label': f"Tháng {month:02d}/{year}", 'value': f"{year}-{month:02d}",
        })

    date_headers = [
        {'date': d, 'label': get_weekday_label(d), 'display': f"{get_weekday_label(d)} {d.strftime('%d/%m')}"}
        for d in dates
    ]

    summary_daily = []
    for i, d in enumerate(dates):
        day_hc = Decimal('0.0')
        day_tc = Decimal('0.0')
        for row in payroll_data:
            att = row['daily_attendance'][i]
            hc_sym = att['hc_symbol']
            tc_sym = att['tc_symbol']
            if hc_sym == 'x':   day_hc += Decimal('1.0')
            elif hc_sym == '/': day_hc += Decimal('0.5')
            if tc_sym:
                try: day_tc += Decimal(str(tc_sym))
                except: pass
        summary_daily.append({'hc': day_hc, 'tc': day_tc})

    summary_total_hc = sum((Decimal(r['total_hc']) for r in payroll_data), Decimal(0)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    summary_total_tc = sum((Decimal(r['total_tc']) for r in payroll_data), Decimal(0)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    summary_total_pay = sum((Decimal(r['total_pay']) for r in payroll_data), Decimal(0))

    context = {
        'target_date': target_date, 'dates': dates, 'date_headers': date_headers,
        'start_date': start_date, 'end_date': end_date,
        'prev_week': start_date - datetime.timedelta(days=7),
        'next_week': start_date + datetime.timedelta(days=7),
        'payroll_data': payroll_data, 'search_query': search_query, 'search_type': search_type,
        'available_weeks': available_weeks, 'available_months': available_months,
        'view_type': view_type, 'current_month': target_date.strftime('%m'),
        'can_edit': can_edit, 'is_viewer_mode': is_viewer_mode,
        'summary_daily': summary_daily, 'summary_total_hc': summary_total_hc,
        'summary_total_tc': summary_total_tc, 'summary_total_pay': summary_total_pay,
        'du_an_id': du_an_id, 'du_an_obj': du_an_obj,
    }
    return render(request, 'payroll/payroll_sheet.html', context)


from django.http import JsonResponse

def get_adjustment_logs(request, employee_id):
    logs = AdjustmentLog.objects.filter(employee_id=employee_id).order_by('-created_at')
    data = []
    for log in logs:
        data.append({
            'ngay': log.created_at.strftime('%d/%m/%Y %H:%M'),
            'tang': float(log.amount) if log.amount > 0 else 0,
            'giam': abs(float(log.amount)) if log.amount < 0 else 0,
        })
    return JsonResponse({'logs': data})


def payroll_statistics(request):
    can_edit = False
    is_viewer_mode = True
    if request.user.is_authenticated:
        try:
            role = request.user.profile.role
            if role in ['manager', 'user']:
                can_edit = True
                is_viewer_mode = False
        except:
            messages.error(request, 'Lỗi: Hồ sơ người dùng không hợp lệ.')
            return redirect('login:login')

    du_an_id = request.GET.get('du_an_id', '').strip()
    du_an_obj = None
    if du_an_id:
        try:
            du_an_obj = CongTrinh.objects.get(id=du_an_id)
        except (CongTrinh.DoesNotExist, ValueError):
            du_an_id = ''

    search_query = request.GET.get('q', '').strip()
    view_type = request.GET.get('view_type', 'week').strip().lower()
    if view_type not in ('week', 'month'):
        view_type = 'week'

    # ✅ FIX: Nếu có dự án và không truyền ?date= → tự động về tuần đầu dự án
    if du_an_obj and 'date' not in request.GET:
        target_date = get_week_start(du_an_obj.ngay_bat_dau)
    else:
        target_date = get_target_date(request)

    if view_type == 'month':
        week_start = get_week_start(target_date)
        year, month = week_start.year, week_start.month
        weeks_this_month = get_weeks_in_month(year, month, du_an_id=du_an_id)
        if weeks_this_month:
            first_day = weeks_this_month[0]
            last_day = weeks_this_month[-1] + datetime.timedelta(days=6)
            dates = [first_day + datetime.timedelta(days=i) for i in range((last_day - first_day).days + 1)]
        else:
            dates = get_week_range(target_date)
    else:
        dates = get_week_range(target_date)

    SUMMARY_NAMES = [
    'Tổng số Cai', 'Tổng số Kho', 'Tổng số LĐ', 
    'Người chấm công', 'Cộng', 'Tổng cộng'
]
    if du_an_id:
        emp_ids_in_project = Attendance.objects.filter(
            cong_trinh_id=du_an_id
        ).values_list('employee_id', flat=True).distinct()
        employees = Employee.objects.filter(
            id__in=emp_ids_in_project
        ).exclude(name__in=SUMMARY_NAMES).order_by('name')
    else:
        employees = Employee.objects.exclude(name__in=SUMMARY_NAMES).order_by('name')

    if search_query:
        employees = employees.filter(name__icontains=search_query)

    start_date, end_date = dates[0], dates[-1]
    payroll_data = build_weekly_payroll_data(employees, dates, du_an_id=du_an_id)

    total_hc = sum((row['total_hc'] for row in payroll_data), Decimal('0.0'))
    total_tc = sum((row['total_tc'] for row in payroll_data), Decimal('0.0'))
    total_pay = sum((row['total_pay'] for row in payroll_data), Decimal('0'))

    daily_totals = []
    for current_date in dates:
        daily_hc = Decimal('0.0')
        daily_tc = Decimal('0.0')
        for row in payroll_data:
            for att in row['daily_attendance']:
                if att['date'] == current_date:
                    daily_hc += from_symbol(att['hc_symbol'])
                    daily_tc += from_symbol(att['tc_symbol'])
                    break
        daily_totals.append({'label': current_date.strftime('%d/%m'), 'hc': float(daily_hc), 'tc': float(daily_tc)})

    chart_rows = [
        {'name': row['employee'].name, 'hc': float(row['total_hc']), 'tc': float(row['total_tc'])}
        for row in payroll_data
    ]

    all_weeks_raw = get_available_weeks(du_an_id=du_an_id)
    available_weeks = [
        {
            'start': w, 'end': w + datetime.timedelta(days=6),
            'date_str': w.strftime('%Y-%m-%d'),
            'label': f"Tuần {i:02d}: {w.strftime('%d/%m')} – {(w + datetime.timedelta(days=6)).strftime('%d/%m/%Y')}",
        }
        for i, w in enumerate(all_weeks_raw, start=1)
    ]
    available_months = []
    for (year, month) in get_available_months(du_an_id=du_an_id):
        weeks_in_m = get_weeks_in_month(year, month, du_an_id=du_an_id)
        first_week_str = weeks_in_m[0].strftime('%Y-%m-%d') if weeks_in_m else f'{year}-{month:02d}-01'
        available_months.append({'year': year, 'month': month, 'date_str': first_week_str, 'label': f"Tháng {month:02d}/{year}"})

    context = {
        'start_date': start_date, 'end_date': end_date, 'target_date': target_date,
        'view_type': view_type,
        'prev_week': start_date - datetime.timedelta(days=7),
        'next_week': start_date + datetime.timedelta(days=7),
        'payroll_data': payroll_data, 'employee_count': len(payroll_data),
        'total_hc': total_hc, 'total_tc': total_tc, 'total_pay': total_pay,
        'daily_totals': daily_totals, 'chart_rows': chart_rows,
        'available_weeks': available_weeks, 'available_months': available_months,
        'search_query': search_query, 'can_edit': can_edit, 'is_viewer_mode': is_viewer_mode,
        'du_an_id': du_an_id, 'du_an_obj': du_an_obj,
    }
    return render(request, 'payroll/payroll_statistics.html', context)

# 
def add_employee(request):
    if not request.user.is_authenticated:
        messages.error(request, 'Bạn phải đăng nhập để thực hiện chức năng này.')
        return redirect('login:login')
    try:
        role = request.user.profile.role
        if role not in ['manager', 'user']:
            messages.error(request, 'Bạn không có quyền thêm nhân viên.')
            return redirect('payroll:payroll_sheet')
    except:
        messages.error(request, 'Lỗi: Hồ sơ người dùng không hợp lệ.')
        return redirect('login:login')

    if request.method == 'POST':
        name = request.POST.get('name')
        date_of_birth = request.POST.get('date_of_birth') or None
        pos = request.POST.get('position')
        wage = request.POST.get('daily_wage', 0)
        du_an_id = request.POST.get('du_an_id', '').strip()
        current_date = request.POST.get('current_date', '').strip()

        emp = Employee.objects.create(
            name=name, date_of_birth=date_of_birth, position=pos, daily_wage=wage
        )
        if du_an_id:
            try:
                cong_trinh_obj = CongTrinh.objects.get(id=du_an_id)

                # ✅ Ưu tiên lấy ngày bắt đầu của dự án
                if cong_trinh_obj.ngay_bat_dau:
                    anchor_date = cong_trinh_obj.ngay_bat_dau
                elif current_date:
                    anchor_date = datetime.datetime.strptime(current_date, '%Y-%m-%d').date()
                else:
                    anchor_date = timezone.now().date()

                week_dates = get_week_range(anchor_date)
                first_day_of_week = week_dates[0]

                Attendance.objects.get_or_create(
                    employee=emp,
                    date=first_day_of_week,
                    cong_trinh=cong_trinh_obj,
                    defaults={
                        'regular_workday': Decimal('0.0'),
                        'overtime_workday': Decimal('0.0'),
                    }
                )
            except (CongTrinh.DoesNotExist, ValueError):
                pass

        messages.success(request, f'Đã thêm nhân viên "{name}" thành công.')

        redirect_url = reverse('payroll:payroll_sheet')
        params = []
        if du_an_id:
            try:
                cong_trinh_obj = CongTrinh.objects.get(id=du_an_id)
                if cong_trinh_obj.ngay_bat_dau:
                    week_start = get_week_start(cong_trinh_obj.ngay_bat_dau)
                    params.append(f'date={week_start.strftime("%Y-%m-%d")}')
                elif current_date:
                    params.append(f'date={current_date}')
            except (CongTrinh.DoesNotExist, ValueError):
                if current_date:
                    params.append(f'date={current_date}')
            params.append(f'du_an_id={du_an_id}')
        elif current_date:
            params.append(f'date={current_date}')

        if params:
            redirect_url += '?' + '&'.join(params)
        return redirect(redirect_url)

    return redirect('payroll:payroll_sheet')
# 

def edit_employee(request, pk):
    if not request.user.is_authenticated:
        messages.error(request, 'Bạn phải đăng nhập để thực hiện chức năng này.')
        return redirect('login:login')
    try:
        role = request.user.profile.role
        if role not in ['manager', 'user']:
            messages.error(request, 'Bạn không có quyền sửa nhân viên.')
            return redirect('payroll:payroll_sheet')
    except:
        messages.error(request, 'Lỗi: Hồ sơ người dùng không hợp lệ.')
        return redirect('login:login')

    emp = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        emp.name = request.POST.get('name')
        emp.date_of_birth = request.POST.get('date_of_birth') or None
        emp.position = request.POST.get('position')
        emp.daily_wage = request.POST.get('daily_wage')
        emp.save()
        log_activity(request, 'UPDATE', 'NhanVien', object_id=str(pk), object_repr=emp.name, description=f'Sửa nhân viên: {emp.name}')
        messages.success(request, f'Đã cập nhật nhân viên "{emp.name}" thành công.')
    return redirect('payroll:payroll_sheet')


def delete_employee(request, pk):
    if not request.user.is_authenticated:
        messages.error(request, 'Bạn phải đăng nhập để thực hiện chức năng này.')
        return redirect('login:login')
    try:
        role = request.user.profile.role
        if role not in ['manager', 'user']:
            messages.error(request, 'Bạn không có quyền xóa nhân viên.')
            return redirect('payroll:payroll_sheet')
    except:
        messages.error(request, 'Lỗi: Hồ sơ người dùng không hợp lệ.')
        return redirect('login:login')

    emp = get_object_or_404(Employee, pk=pk)
    emp_name = emp.name
    emp.delete()
    log_activity(request, 'DELETE', 'NhanVien', object_repr=emp_name, description=f'Xóa nhân viên: {emp_name}')
    messages.success(request, f'Đã xóa nhân viên "{emp_name}" thành công.')
    return redirect('payroll:payroll_sheet')


def import_excel(request):
    if not request.user.is_authenticated:
        messages.error(request, 'Bạn phải đăng nhập để thực hiện chức năng này.')
        return redirect('login:login')
    try:
        role = request.user.profile.role
        if role != 'manager':
            messages.error(request, 'Chỉ quản lý mới có thể nhập file Excel.')
            return redirect('payroll:payroll_sheet')
    except:
        messages.error(request, 'Lỗi: Hồ sơ người dùng không hợp lệ.')
        return redirect('login:login')

    du_an_id = request.POST.get('du_an_id', '').strip()
    cong_trinh_obj = None
    # Thêm vào đầu import_excel, sau khi xác định cong_trinh_obj
    if du_an_id and not cong_trinh_obj:
        messages.warning(request, f'Không tìm thấy dự án ID={du_an_id}. Dữ liệu sẽ không gắn vào dự án nào.')
    if du_an_id:
        try:
            cong_trinh_obj = CongTrinh.objects.get(id=du_an_id)
        except (CongTrinh.DoesNotExist, ValueError):
            cong_trinh_obj = None

    first_target_date = None

    if request.method == 'POST' and request.FILES.get('excel_file'):
        from openpyxl import load_workbook
        import re

        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, data_only=True)

        for ws in wb.worksheets:
            
            sheet_name = ws.title.strip().lower()
            if 'tổng' in sheet_name or 'tong' in sheet_name:
                continue
            header_text = str(ws['G1'].value or '')
            date_matches = re.findall(r'(\d{1,2}/\d{1,2}/(?:\d{4}|\d{2}))', header_text)
            if not date_matches:
                continue
            raw_date = date_matches[0]
            try:
                try:
                    target_date = datetime.datetime.strptime(raw_date, '%d/%m/%Y').date()
                except ValueError:
                    target_date = datetime.datetime.strptime(raw_date, '%d/%m/%y').date()
            except Exception:
                continue
            if first_target_date is None:
                first_target_date = target_date
            dates = get_week_range(target_date)
            SKIP_NAMES = [
                'Tổng số Cai', 'Tổng số Kho', 'Tổng số LĐ', 
                'Người chấm công', 'Cộng', 'Tổng cộng', 'STT'
            ]
            for row in ws.iter_rows(min_row=6):
                name = row[1].value
                if not name:
                    continue
                name = str(name).strip()
                if name in SKIP_NAMES:
                    continue
                position = row[2].value
                emp, _ = Employee.objects.get_or_create(
                    name=name, defaults={'position': position or 'Thợ', 'daily_wage': 0}
                )
                if position and emp.position != position:
                    emp.position = position
                    emp.save()
                col_idx = 3
                for d in dates:
                    hc_val = from_symbol(row[col_idx].value)
                    tc_raw = row[col_idx + 1].value
                    try:
                        tc_val = Decimal(str(tc_raw)) if tc_raw else Decimal('0')
                    except:
                        tc_val = Decimal('0')
                    att, _ = Attendance.objects.get_or_create(
                        employee=emp, date=d, cong_trinh=cong_trinh_obj
                    )
                    att.regular_workday = hc_val
                    att.overtime_workday = tc_val
                    att.save()
                    col_idx += 2
                wage = row[19].value
                if wage:
                    try:
                        emp.daily_wage = Decimal(str(wage).replace('.', '').replace(',', ''))
                    except:
                        pass
                    emp.save()
                adj_val_raw = row[20].value
                if adj_val_raw:
                    try:
                        adj_val = Decimal(str(adj_val_raw).replace('.', '').replace(',', ''))
                    except:
                        adj_val = Decimal('0')
                    if adj_val != Decimal('0'):
                        adj, _ = Adjustment.objects.get_or_create(
                            employee=emp, start_date=dates[0], end_date=dates[-1]
                        )
                        adj.amount = adj_val
                        adj.save()
                        AdjustmentLog.objects.create(
                            employee=emp, start_date=dates[0], end_date=dates[-1], amount=adj_val
                        )

    redirect_date = first_target_date.strftime('%Y-%m-%d') if first_target_date else request.POST.get('current_date', '')
    redirect_url = reverse('payroll:payroll_sheet') + f"?date={redirect_date}"
    if du_an_id:
        redirect_url += f"&du_an_id={du_an_id}"
    return redirect(redirect_url)


def delete_all_data(request):
    if not request.user.is_authenticated:
        messages.error(request, 'Bạn phải đăng nhập để thực hiện chức năng này.')
        return redirect('login:login')
    try:
        role = request.user.profile.role
        if role != 'manager':
            messages.error(request, 'Chỉ quản lý mới có thể xóa dữ liệu.')
            return redirect('payroll:payroll_sheet')
    except:
        messages.error(request, 'Lỗi: Hồ sơ người dùng không hợp lệ.')
        return redirect('login:login')

    du_an_id = request.POST.get('du_an_id', '').strip()
    if du_an_id:
        try:
            du_an = CongTrinh.objects.get(id=du_an_id)
            deleted_count, _ = Attendance.objects.filter(cong_trinh=du_an).delete()
            messages.success(request, f'Đã xóa {deleted_count} bản ghi chấm công của dự án "{du_an.ten_cong_trinh}".')
        except CongTrinh.DoesNotExist:
            messages.error(request, 'Dự án không tồn tại.')
    else:
        Employee.objects.all().delete()
        messages.success(request, 'Đã xóa tất cả dữ liệu nhân viên.')

    redirect_url = reverse('payroll:payroll_sheet')
    if du_an_id:
        redirect_url += f'?du_an_id={du_an_id}'
    return redirect(redirect_url)


def save_attendance(request):
    if not request.user.is_authenticated:
        messages.error(request, 'Bạn phải đăng nhập để thực hiện chức năng này.')
        return redirect('login:login')
    try:
        role = request.user.profile.role
        if role not in ['manager', 'user']:
            messages.error(request, 'Bạn không có quyền lưu dữ liệu chấm công.')
            return redirect('payroll:payroll_sheet')
    except:
        messages.error(request, 'Lỗi: Hồ sơ người dùng không hợp lệ.')
        return redirect('login:login')

    if request.method == 'POST':
        current_date_str = request.POST.get('current_date')
        du_an_id = request.POST.get('du_an_id', '').strip()
        cong_trinh_obj = None
        if du_an_id:
            try:
                cong_trinh_obj = CongTrinh.objects.get(id=du_an_id)
            except (CongTrinh.DoesNotExist, ValueError):
                du_an_id = ''

        try:
            target_date = datetime.datetime.strptime(current_date_str, '%Y-%m-%d').date()
        except Exception:
            target_date = timezone.now().date()

        view_type = request.POST.get('view_type', 'week')
        if view_type == 'month':
            first_day = target_date.replace(day=1)
            last_day = target_date.replace(day=calendar.monthrange(target_date.year, target_date.month)[1])
            dates = [first_day + datetime.timedelta(days=i) for i in range((last_day - first_day).days + 1)]
        else:
            dates = get_week_range(target_date)

        start_date, end_date = dates[0], dates[-1]
        search_query = request.POST.get('q', '').strip()

        if du_an_id and cong_trinh_obj:
            emp_ids_in_project = Attendance.objects.filter(
                cong_trinh=cong_trinh_obj
            ).values_list('employee_id', flat=True).distinct()
            employees = Employee.objects.filter(id__in=emp_ids_in_project).order_by('name')
            if search_query:
                employees = employees.filter(name__icontains=search_query)
        else:
            employees = get_filtered_employees(search_query)

        for emp in employees:
            for d in dates:
                date_str = d.strftime('%Y-%m-%d')
                hc_val = parse_attendance_input(request.POST.get(f'att_{emp.id}_{date_str}_hc'))
                tc_val = parse_attendance_input(request.POST.get(f'att_{emp.id}_{date_str}_tc'))

                if hc_val == Decimal('0.0') and tc_val == Decimal('0.0'):
                    if cong_trinh_obj:
                        att = Attendance.objects.filter(
                            employee=emp, date=d, cong_trinh=cong_trinh_obj
                        ).first()
                        if att:
                            att.regular_workday = Decimal('0.0')
                            att.overtime_workday = Decimal('0.0')
                            att.save()
                    else:
                        Attendance.objects.filter(employee=emp, date=d, cong_trinh=None).delete()
                else:
                    att, _ = Attendance.objects.get_or_create(
                        employee=emp, date=d, cong_trinh=cong_trinh_obj
                    )
                    att.regular_workday = hc_val
                    att.overtime_workday = tc_val
                    att.save()

            adj_name = f'adj_{emp.id}_{start_date.strftime("%Y-%m-%d")}_{end_date.strftime("%Y-%m-%d")}'
            adj_val = parse_attendance_input(request.POST.get(adj_name))
            adjustment, created = Adjustment.objects.get_or_create(
                employee=emp, start_date=start_date, end_date=end_date,
                defaults={'amount': Decimal('0.0')}
            )
            if adj_val == Decimal('0.0'):
                if not created:
                    adjustment.delete()
            else:
                adjustment.amount = adj_val
                adjustment.save()

        messages.success(request, 'Đã lưu dữ liệu chấm công và các khoản tăng/giảm thành công.')
        redirect_url = reverse('payroll:payroll_sheet') + f"?date={current_date_str}&view_type={view_type}"
        if search_query:
            redirect_url += f"&q={search_query}"
        if du_an_id:
            redirect_url += f"&du_an_id={du_an_id}"
        return redirect(redirect_url)

    return redirect('payroll:payroll_sheet')


def export_payroll_excel(dates, start_date, end_date, view_type='week', du_an_id=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bang Cham Cong"
    ws.merge_cells('A1:E1')
    ws['A1'] = "CÔNG TY TNHH XÂY DỰNG CPT"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:E2')
    ws['A2'] = "Địa chỉ: TP. Hồ Chí Minh"
    total_cols_count = 3 + 2 * len(dates) + 7
    from openpyxl.utils import get_column_letter
    max_col_letter = get_column_letter(total_cols_count)
    ws.merge_cells(f'A3:{max_col_letter}3')
    if view_type == 'month':
        ws['A3'] = f"BẢNG CHẤM CÔNG VÀ TÍNH LƯƠNG (Tháng: {start_date.strftime('%m/%Y')})"
    else:
        ws['A3'] = f"BẢNG CHẤM CÔNG VÀ TÍNH LƯƠNG (Tuần: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})"
    ws['A3'].font = Font(bold=True, size=14)
    ws['A3'].alignment = Alignment(horizontal='center')
    headers = ['STT', 'Họ tên', 'Nghề']
    ws.append(headers)
    for col in ['A', 'B', 'C']:
        ws.merge_cells(f'{col}4:{col}5')
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}4'].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    ws['A4'], ws['B4'], ws['C4'] = 'STT', 'Họ tên', 'Nghề'
    col_offset = 4
    v_days = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]
    for i, d in enumerate(dates):
        cell = ws.cell(row=4, column=col_offset)
        weekday_idx = d.weekday()
        cell.value = f"{v_days[weekday_idx]} {d.strftime('%d/%m')}"
        ws.merge_cells(start_row=4, start_column=col_offset, end_row=4, end_column=col_offset+1)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
        ws.cell(row=5, column=col_offset).value = "HC"
        ws.cell(row=5, column=col_offset+1).value = "TC"
        ws.cell(row=5, column=col_offset).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        col_offset += 2
    res_headers = ['Tổng HC', 'Tổng TC', 'Lương ngày', 'Tăng/Giảm', 'Tổng lãnh', 'Ký nhận', 'Ghi chú']
    for i, h in enumerate(res_headers):
        cell = ws.cell(row=4, column=col_offset + i)
        cell.value = h
        ws.merge_cells(start_row=4, start_column=col_offset + i, end_row=5, end_column=col_offset + i)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    SUMMARY_NAMES = [
        'Tổng số Cai', 'Tổng số Kho', 'Tổng số LĐ', 
        'Người chấm công', 'Cộng', 'Tổng cộng'
    ]
    if du_an_id:
        emp_ids = Attendance.objects.filter(
            cong_trinh_id=du_an_id
        ).values_list('employee_id', flat=True).distinct()
        employees = Employee.objects.filter(id__in=emp_ids).exclude(name__in=SUMMARY_NAMES).order_by('name')
    else:
        employees = Employee.objects.exclude(name__in=SUMMARY_NAMES).order_by('name')
    row_num = 6
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for idx, emp in enumerate(employees, 1):
        ws.cell(row=row_num, column=1).value = idx
        ws.cell(row=row_num, column=2).value = emp.name
        ws.cell(row=row_num, column=3).value = emp.position
        total_hc = Decimal('0.0')
        total_tc = Decimal('0.0')
        c_idx = 4
        for d in dates:
            att_qs = Attendance.objects.filter(employee=emp, date=d)
            if du_an_id:
                att_qs = att_qs.filter(cong_trinh_id=du_an_id)
            att = att_qs.first()
            hc = att.regular_workday if att else Decimal('0.0')
            tc = att.overtime_workday if att else Decimal('0.0')
            ws.cell(row=row_num, column=c_idx).value = to_symbol(hc) if hc > 0 else ""
            ws.cell(row=row_num, column=c_idx+1).value = float(tc) if tc > 0 else ""
            total_hc += hc
            total_tc += tc
            c_idx += 2
        ws.cell(row=row_num, column=c_idx).value = float(total_hc)
        ws.cell(row=row_num, column=c_idx+1).value = float(total_tc)
        ws.cell(row=row_num, column=c_idx+2).value = float(emp.daily_wage)
        ws.cell(row=row_num, column=c_idx+2).number_format = '#,##0'
        adj = Adjustment.objects.filter(employee=emp, start_date=start_date, end_date=end_date).first()
        adj_val = adj.amount if adj else Decimal('0.0')
        ws.cell(row=row_num, column=c_idx+3).value = float(adj_val)
        ws.cell(row=row_num, column=c_idx+3).font = Font(color="FF0000")
        total_pay = (total_hc * emp.daily_wage) + (total_tc * (emp.daily_wage / Decimal('8'))) + adj_val
        ws.cell(row=row_num, column=c_idx+4).value = float(total_pay)
        ws.cell(row=row_num, column=c_idx+4).font = Font(bold=True)
        ws.cell(row=row_num, column=c_idx+4).number_format = '#,##0'
        row_num += 1
    for r in ws.iter_rows(min_row=4, max_row=row_num-1, min_col=1, max_col=col_offset+len(res_headers)-1):
        for cell in r:
            cell.border = border
            if not cell.alignment.horizontal:
                cell.alignment = Alignment(horizontal='center')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if view_type == 'month':
        response['Content-Disposition'] = f'attachment; filename=BangLuongThang_{start_date.strftime("%m_%Y")}.xlsx'
    else:
        response['Content-Disposition'] = f'attachment; filename=BangLuongTuan_{start_date}.xlsx'
    wb.save(response)
    return response


# ============ DỰ ÁN QUẢN LÝ (PROJECT MANAGEMENT) ============

@allow_viewer
def du_an_list(request):
    search_query = request.GET.get('search', '').strip()
    time_filter = request.GET.get('time_filter', 'all')
    status_filter = request.GET.get('status', '')
    du_an = CongTrinh.objects.all()
    if search_query:
        du_an = du_an.filter(
            Q(ten_cong_trinh__icontains=search_query) |
            Q(dia_diem__icontains=search_query)
        )
    if status_filter:
        du_an = du_an.filter(trang_thai=status_filter)
    now = timezone.now().date()
    if time_filter == 'ngay':
        du_an = du_an.filter(ngay_bat_dau=now)
    elif time_filter == 'tuan':
        start_week = now - timedelta(days=now.weekday())
        end_week = start_week + timedelta(days=6)
        du_an = du_an.filter(ngay_bat_dau__gte=start_week, ngay_bat_dau__lte=end_week)
    elif time_filter == 'thang':
        du_an = du_an.filter(ngay_bat_dau__year=now.year, ngay_bat_dau__month=now.month)
    elif time_filter == 'nam':
        du_an = du_an.filter(ngay_bat_dau__year=now.year)
    du_an = du_an.order_by('-ngay_bat_dau')
    context = {
        'du_an': du_an, 'search_query': search_query, 'time_filter': time_filter,
        'status_filter': status_filter, 'status_choices': CongTrinh.TRANG_THAI_CHOICES, 'can_edit': True,
    }
    return render(request, 'payroll/du_an_list.html', context)

@manager_only
def du_an_create(request):
    if request.method == 'POST':
        form = CongTrinhForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thêm công trình mới thành công.')
            return redirect('payroll:du_an_list')
    else:
        form = CongTrinhForm()
    return render(request, 'payroll/du_an_form.html', {'form': form, 'title': 'Thêm công trình'})


@allow_viewer
def du_an_detail(request, pk):
    du_an = get_object_or_404(CongTrinh, pk=pk)
    from .models import Attendance
    SUMMARY_NAMES = ['Tổng số Cai', 'Tổng số Kho', 'Tổng số LĐ', 'Người chấm công']
    nhan_vien_count = Attendance.objects.filter(
        cong_trinh=du_an
    ).exclude(
        employee__name__in=SUMMARY_NAMES
    ).values('employee').distinct().count()

    cham_cong_count = Attendance.objects.filter(
        cong_trinh=du_an
    ).exclude(
        regular_workday=0,
        overtime_workday=0
    ).count()

    # ✅ FIX: get_available_weeks giờ đã sinh tuần từ khoảng thời gian dự án,
    # nên all_weeks luôn có dữ liệu ngay cả khi chưa chấm công lần nào
    all_weeks = get_available_weeks(du_an_id=pk)

    if all_weeks:
        # Dùng tuần đầu tiên của dự án làm điểm mặc định
        latest_week = all_weeks[0].strftime('%Y-%m-%d')
    else:
        # Fallback an toàn: tuần chứa ngày bắt đầu dự án
        latest_week = get_week_start(du_an.ngay_bat_dau).strftime('%Y-%m-%d')

    return render(request, 'payroll/du_an_detail.html', {
        'du_an': du_an,
        'cham_cong_count': cham_cong_count,
        'nhan_vien_count': nhan_vien_count,
        'latest_week': latest_week,
    })

@manager_only
def du_an_edit(request, pk):
    du_an = get_object_or_404(CongTrinh, pk=pk)
    if request.method == 'POST':
        form = CongTrinhForm(request.POST, instance=du_an)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cập nhật công trình "{du_an.ten_cong_trinh}" thành công.')
            return redirect('payroll:du_an_list')
    else:
        form = CongTrinhForm(instance=du_an)
    return render(request, 'payroll/du_an_form.html', {'form': form, 'title': 'Sửa công trình'})

@manager_only
def du_an_delete(request, pk):
    du_an = get_object_or_404(CongTrinh, pk=pk)
    if request.method == 'POST':
        ten = du_an.ten_cong_trinh
        du_an.delete()
        messages.success(request, f'Đã xóa công trình "{ten}".')
        return redirect('payroll:du_an_list')
    return render(request, 'payroll/du_an_confirm_delete.html', {'du_an': du_an})

@allow_viewer
def tong_hop_2026(request):
    from itertools import groupby
    all_weeks = get_available_weeks()
    if not all_weeks:
        messages.error(request, 'Chưa có dữ liệu. Vui lòng import file Excel trước.')
        return redirect('payroll:payroll_statistics')
    SUMMARY_NAMES = [
    'Tổng số Cai', 'Tổng số Kho', 'Tổng số LĐ', 
    'Người chấm công', 'Cộng', 'Tổng cộng'
]
    employees = Employee.objects.exclude(name__in=SUMMARY_NAMES).order_by('name')
    tuan_cols = [
        {
            'col': i, 'label': f'T{i+1:02d}', 'date': f"{w.strftime('%d/%m')}",
            'week_start': w, 'week_end': w + datetime.timedelta(days=6),
        }
        for i, w in enumerate(all_weeks)
    ]
    start = all_weeks[0]
    end = all_weeks[-1] + datetime.timedelta(days=6)
    all_atts = Attendance.objects.filter(date__range=(start, end)).select_related('employee')
    att_map = {}
    for att in all_atts:
        att_map.setdefault(att.employee_id, {})[att.date] = att
    all_adjs = Adjustment.objects.filter(start_date__gte=start, end_date__lte=end).select_related('employee')
    adj_map = {}
    for adj in all_adjs:
        adj_map[(adj.employee_id, adj.start_date)] = adj.amount
    employee_rows = []
    col_totals = [0.0] * len(tuan_cols)
    grand_cong = 0.0
    grand_luong = 0.0
    for emp in employees:
        emp_atts = att_map.get(emp.id, {})
        week_vals = []
        row_cong = 0.0
        for idx, t in enumerate(tuan_cols):
            week_cong = Decimal('0.0')
            d = t['week_start']
            while d <= t['week_end']:
                att = emp_atts.get(d)
                if att:
                    week_cong += att.regular_workday
                d += datetime.timedelta(days=1)
            val = int(float(week_cong) + 0.5) if week_cong > 0 else None
            week_vals.append(val)
            if val:
                row_cong += val
                col_totals[idx] += val
        tong_luong_nv = row_cong * float(emp.daily_wage)
        row_cong = round(row_cong, 1)
        grand_cong += row_cong
        grand_luong += tong_luong_nv
        employee_rows.append({
            'ten': emp.name, 'nghe': emp.position, 'luong_ngay': int(emp.daily_wage),
            'week_vals': week_vals, 'tong_cong': row_cong, 'tong_luong': tong_luong_nv,
        })
    employee_rows = [e for e in employee_rows if e['tong_cong'] > 0]
    context = {
        'ten_cty': 'CTY CP XÂY DỰNG CPT', 'ten_bang': 'BẢNG TỔNG HỢP NGÀY CÔNG',
        'khoang_tg': f"{all_weeks[0].strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}",
        'ten_ct': 'NHÀ PHỐ - A.LÂM - CỒN KHƯƠNG',
        'tuan_cols': tuan_cols, 'employees': employee_rows, 'col_totals': col_totals,
        'grand_cong': grand_cong, 'grand_luong': grand_luong, 'so_nv': len(employee_rows),
    }
    return render(request, 'payroll/tong_hop_2026.html', context)

# ── 1. Dashboard tổng quan theo dõi user ─────────────────
@login_required
@manager_only
def user_tracking_dashboard(request):
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
 
    # ---- Thống kê nhanh ----
    online_cutoff = now - timezone.timedelta(minutes=5)
    online_users  = ActiveSession.objects.filter(
        last_activity__gte=online_cutoff
    ).select_related('user').order_by('-last_activity')
 
    logins_today = UserLoginLog.objects.filter(
        action=UserLoginLog.ACTION_LOGIN,
        created_at__gte=today_start
    ).count()
 
    failed_today = UserLoginLog.objects.filter(
        action=UserLoginLog.ACTION_FAILED,
        created_at__gte=today_start
    ).count()
 
    # ---- Activity log gần nhất (50 dòng) ----
    recent_activities = UserActivityLog.objects.select_related('user') \
                            .order_by('-created_at')[:50]
 
    # ---- Login log gần nhất (50 dòng) ----
    recent_logins = UserLoginLog.objects.select_related('user') \
                        .order_by('-created_at')[:50]
 
    # ---- Top user hoạt động hôm nay ----
    top_users = UserActivityLog.objects.filter(
        created_at__gte=today_start
    ).values('user__username').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
 
    # ---- Thống kê theo loại hành động hôm nay ----
    action_stats = UserActivityLog.objects.filter(
        created_at__gte=today_start
    ).values('action_type').annotate(total=Count('id')).order_by('-total')
 
    context = {
        'online_users':      online_users,
        'online_count':      online_users.count(),
        'logins_today':      logins_today,
        'failed_today':      failed_today,
        'recent_activities': recent_activities,
        'recent_logins':     recent_logins,
        'top_users':         top_users,
        'action_stats':      action_stats,
        'now':               now,
    }
    return render(request, 'payroll/user_tracking_dashboard.html', context)
 
 
# ── 2. Lịch sử của 1 user cụ thể ─────────────────────────
@login_required
@manager_only
def user_tracking_detail(request, user_id):
    target_user = get_object_or_404(User, pk=user_id)
 
    activities = UserActivityLog.objects.filter(
        user=target_user
    ).order_by('-created_at')[:200]
 
    logins = UserLoginLog.objects.filter(
        user=target_user
    ).order_by('-created_at')[:100]
 
    # Thống kê tổng hợp
    total_actions = activities.count()
    action_breakdown = UserActivityLog.objects.filter(
        user=target_user
    ).values('action_type').annotate(total=Count('id')).order_by('-total')
 
    context = {
        'target_user':       target_user,
        'activities':        activities,
        'logins':            logins,
        'total_actions':     total_actions,
        'action_breakdown':  action_breakdown,
    }
    return render(request, 'payroll/user_tracking_detail.html', context)
 
 
# ── 3. API endpoint — refresh danh sách online (AJAX) ────
@login_required
@manager_only
def api_online_users(request):
    cutoff = timezone.now() - timezone.timedelta(minutes=5)
    sessions = ActiveSession.objects.filter(
        last_activity__gte=cutoff
    ).select_related('user')
 
    data = []
    for s in sessions:
        data.append({
            'username':      s.user.username,
            'full_name':     s.user.get_full_name() or s.user.username,
            'current_page':  s.current_page,
            'last_activity': s.last_activity.strftime('%H:%M:%S'),
            'ip':            s.ip_address or '—',
            'duration':      s.duration,
        })
 
    return JsonResponse({'users': data, 'count': len(data)})
 